import openpyxl

archivo_excel = openpyxl.load_workbook(r"Prueba\Excel_archivos\notas.xlsx")
notas_sheet = archivo_excel["Notas"]

lista_estudiantes = {}

# Nos da como resultado el número de filas
# notas.max_row 

for estudiante_row in range(2, notas_sheet.max_row + 1):
    nombre = notas_sheet.cell(estudiante_row,1).value
    nota = notas_sheet.cell(estudiante_row,2).value
    lista_estudiantes[nombre] = nota

acu = 0
for nota in lista_estudiantes.values():
    acu += nota

n_estudiantes = notas_sheet.max_row-1
promedio_notas = acu/n_estudiantes
promedio = notas_sheet.cell(6,3)
promedio.value = promedio_notas

archivo_excel.save(r"Prueba\Excel_archivos\resultados.xlsx")